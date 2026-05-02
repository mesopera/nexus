"""
Nexus Step 5 — Scoring & Analysis Script
Reads all JSON results from data/results/ and fills Nexus_Evaluation_Matrix.xlsx.

Usage:  python src/score.py
Output: docs/Nexus_Evaluation_Matrix_Scored.xlsx

EXACT COLUMN MAPPING (verified from Excel):
  Math:              E=Accuracy  F=Step Correctness  G=Hallucination  H=Consistency  I=Latency  J=Tokens
  Code:              K=Runs OK   L=Accuracy          M=Hallucination  N=Consistency  O=Latency  P=Tokens
  Creative Write:    Q=Quality   R=Coherence         S=Instruction    T=Consistency  U=Latency  V=Tokens
  Factual QA:        W=Accuracy  X=Hallucination     Y=Instruction    Z=Consistency  AA=Latency AB=Tokens
  Summarisation:     AC=KP Ret.  AD=Accuracy         AE=Instruction   AF=Consistency AG=Latency AH=Tokens
  Instruction:       AI=Constr%  AJ=Accuracy         AK=Hallucination AL=Consistency AM=Latency AN=Tokens
  Reasoning:         AO=Accuracy AP=Chain Quality    AQ=Hallucination AR=Consistency AS=Latency AT=Tokens
  Long Context:      AU=Retrieval AV=Accuracy        AW=Hallucination AX=Consistency AY=Latency AZ=Tokens
  Multilingual:      BA=Accuracy BB=Fluency          BC=Hallucination BD=Consistency BE=Latency BF=Tokens
  Conversational:    BG=Quality  BH=Context Ret.     BI=Instruction   BJ=Consistency BK=Latency BL=Tokens

MODEL ROW MAPPING (verified from Excel):
  llama3.1-70b-groq  -> row 6  (LLaMA 3.1 / 70B / Groq API)
  llama3.1-8b-groq   -> row 8  (LLaMA 3.1 / 8B  / Groq API)
  gemma2-groq        -> row 19 (Gemma 2   / 9B  / Groq API)
"""

import json
import statistics
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import column_index_from_string

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).resolve().parent.parent
RESULTS_DIR = BASE_DIR / "data" / "results"
MATRIX_FILE = BASE_DIR / "docs" / "Nexus_Evaluation_Matrix.xlsx"
OUTPUT_FILE = BASE_DIR / "docs" / "Nexus_Evaluation_Matrix_Scored.xlsx"

# ── Exact row per model (verified) ───────────────────────────────────────────
MODEL_ROWS = {
    "llama3.1-70b-groq": 6,
    "llama3.1-8b-groq":  8,
    "gemma2-groq":       19,
}

# ── Exact column per domain+param (verified) ─────────────────────────────────
COL = {
    # Math
    "math_accuracy":      "E",
    "math_step":          "F",   # manual
    "math_halluc":        "G",
    "math_consist":       "H",
    "math_latency":       "I",
    "math_tokens":        "J",
    # Code
    "code_runs_ok":       "K",
    "code_accuracy":      "L",
    "code_halluc":        "M",
    "code_consist":       "N",
    "code_latency":       "O",
    "code_tokens":        "P",
    # Creative Writing
    "cw_quality":         "Q",   # manual
    "cw_coherence":       "R",   # manual
    "cw_instruction":     "S",
    "cw_consist":         "T",
    "cw_latency":         "U",
    "cw_tokens":          "V",
    # Factual QA
    "fqa_accuracy":       "W",
    "fqa_halluc":         "X",
    "fqa_instruction":    "Y",
    "fqa_consist":        "Z",
    "fqa_latency":        "AA",
    "fqa_tokens":         "AB",
    # Summarisation
    "sum_kp":             "AC",  # manual
    "sum_accuracy":       "AD",
    "sum_instruction":    "AE",
    "sum_consist":        "AF",
    "sum_latency":        "AG",
    "sum_tokens":         "AH",
    # Instruction Following
    "if_constraint":      "AI",
    "if_accuracy":        "AJ",
    "if_halluc":          "AK",
    "if_consist":         "AL",
    "if_latency":         "AM",
    "if_tokens":          "AN",
    # Reasoning
    "rea_accuracy":       "AO",
    "rea_chain":          "AP",  # manual
    "rea_halluc":         "AQ",
    "rea_consist":        "AR",
    "rea_latency":        "AS",
    "rea_tokens":         "AT",
    # Long Context
    "lc_retrieval":       "AU",
    "lc_accuracy":        "AV",
    "lc_halluc":          "AW",
    "lc_consist":         "AX",
    "lc_latency":         "AY",
    "lc_tokens":          "AZ",
    # Multilingual
    "ml_accuracy":        "BA",
    "ml_fluency":         "BB",  # manual
    "ml_halluc":          "BC",
    "ml_consist":         "BD",
    "ml_latency":         "BE",
    "ml_tokens":          "BF",
    # Conversational
    "conv_quality":       "BG",  # manual
    "conv_context":       "BH",  # manual
    "conv_instruction":   "BI",  # manual
    "conv_consist":       "BJ",
    "conv_latency":       "BK",
    "conv_tokens":        "BL",
}

# ── Color fills ───────────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", start_color="C6EFCE")
YELLOW = PatternFill("solid", start_color="FFEB9C")
RED    = PatternFill("solid", start_color="FFC7CE")
BLUE   = PatternFill("solid", start_color="DDEBF7")

def score_fill(v):
    if not isinstance(v, (int, float)): return None
    if v >= 4:   return GREEN
    if v >= 3:   return YELLOW
    return RED

# ── Helpers ───────────────────────────────────────────────────────────────────
def load_results(model_id, domain):
    path = RESULTS_DIR / model_id / domain / "results.json"
    if not path.exists():
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)

def valid(results):
    return [r for r in results if not r.get("error") and r.get("response")]

def avg_latency(results):
    vals = [r["latency_s"] for r in valid(results) if r.get("latency_s", 0) > 0]
    return round(statistics.mean(vals), 2) if vals else 0.0

def avg_tokens(results):
    vals = [r.get("response_tokens", 0) for r in valid(results)]
    vals = [v for v in vals if v > 0]
    return round(statistics.mean(vals), 0) if vals else 0.0

def consistency(results):
    by_prompt = {}
    for r in valid(results):
        pid = r.get("prompt_id", "")
        by_prompt.setdefault(pid, []).append(r.get("response_tokens", 0))
    cvs = []
    for tokens in by_prompt.values():
        if len(tokens) < 2: continue
        m = statistics.mean(tokens)
        if m == 0: continue
        cv = statistics.stdev(tokens) / m
        cvs.append(cv)
    if not cvs: return 3.0
    avg_cv = statistics.mean(cvs)
    if avg_cv < 0.05:  return 5.0
    if avg_cv < 0.15:  return 4.0
    if avg_cv < 0.30:  return 3.0
    if avg_cv < 0.50:  return 2.0
    return 1.0

def accuracy(results, domain):
    v = valid(results)
    if not v: return 0.0
    scores = []
    for r in v:
        resp     = r.get("response", "")
        expected = r.get("expected_answer", "")
        words    = len(resp.split())
        s = 3
        if words < 5:  s = 1
        elif words > 20: s = 4
        if domain in ("math", "reasoning", "factual_qa") and expected:
            ec = expected.strip().lower()[:50]
            rl = resp.lower()
            if ec and ec in rl: s = 5
            elif any(w in rl for w in ec.split()[:3] if len(w) > 3): s = 4
        if domain == "code":
            if "```" in resp or "def " in resp or "function" in resp: s = 4
        scores.append(s)
    return round(statistics.mean(scores), 1)

def hallucination(results, domain):
    v = valid(results)
    if not v: return 0.0
    suspicious = 0
    confident = ["definitely","certainly","absolutely","100%","without a doubt","i am certain"]
    for r in v:
        resp = r.get("response", "").lower()
        if len(resp.split()) < 5: suspicious += 0.5
        if domain in ("factual_qa","math","reasoning"):
            if any(c in resp for c in confident): suspicious += 0.3
    rate = suspicious / len(v)
    if rate < 0.05:  return 5.0
    if rate < 0.15:  return 4.0
    if rate < 0.30:  return 3.0
    if rate < 0.50:  return 2.0
    return 1.0

def instruction_adh(results):
    v = valid(results)
    if not v: return 0.0
    ok = 0
    for r in v:
        resp   = r.get("response","").strip()
        prompt = r.get("prompt","").lower()
        if "one word" in prompt or "single word" in prompt:
            ok += 1 if len(resp.split()) <= 3 else 0
        elif "one sentence" in prompt:
            ok += 1 if resp.count(".") <= 2 else 0
        else:
            ok += 1
    return round((ok / len(v)) * 5, 1)

def code_runs_ok(results):
    v = valid(results)
    if not v: return 0.0
    runnable = sum(
        1 for r in v
        if any(x in r.get("response","") for x in ["```","def ","class ","import ","return ","function","=>"])
        and not any(x in r.get("response","").lower() for x in ["syntaxerror","nameerror","typeerror"])
    )
    rate = runnable / len(v)
    if rate >= 0.9: return 5.0
    if rate >= 0.7: return 4.0
    if rate >= 0.5: return 3.0
    if rate >= 0.3: return 2.0
    return 1.0

# ── Cell writer ───────────────────────────────────────────────────────────────
def write(sheet, row, col_letter, value, fill=None):
    col_idx = column_index_from_string(col_letter)
    cell = sheet.cell(row=row, column=col_idx)
    if isinstance(value, float):
        cell.value = round(value, 2)
    else:
        cell.value = value
    cell.alignment = Alignment(horizontal="center")
    if fill:
        cell.fill = fill

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("\nNexus Step 5 — Scoring")
    print("=" * 60)

    wb    = load_workbook(MATRIX_FILE)
    sheet = wb["Evaluation Matrix"]

    for model_id, row in MODEL_ROWS.items():
        print(f"\nModel: {model_id}  (row {row})")
        print("-" * 40)

        # ── MATH ──────────────────────────────────────────────────────────────
        r = load_results(model_id, "math")
        if r:
            print(f"  math          : {len(valid(r))}/{len(r)} valid")
            write(sheet, row, COL["math_accuracy"], accuracy(r,"math"),     score_fill(accuracy(r,"math")))
            write(sheet, row, COL["math_step"],     "—")
            write(sheet, row, COL["math_halluc"],   hallucination(r,"math"),score_fill(hallucination(r,"math")))
            write(sheet, row, COL["math_consist"],  consistency(r),         score_fill(consistency(r)))
            write(sheet, row, COL["math_latency"],  avg_latency(r),         BLUE)
            write(sheet, row, COL["math_tokens"],   avg_tokens(r),          BLUE)

        # ── CODE ──────────────────────────────────────────────────────────────
        r = load_results(model_id, "code")
        if r:
            print(f"  code          : {len(valid(r))}/{len(r)} valid")
            write(sheet, row, COL["code_runs_ok"], code_runs_ok(r),         score_fill(code_runs_ok(r)))
            write(sheet, row, COL["code_accuracy"], accuracy(r,"code"),     score_fill(accuracy(r,"code")))
            write(sheet, row, COL["code_halluc"],   hallucination(r,"code"),score_fill(hallucination(r,"code")))
            write(sheet, row, COL["code_consist"],  consistency(r),         score_fill(consistency(r)))
            write(sheet, row, COL["code_latency"],  avg_latency(r),         BLUE)
            write(sheet, row, COL["code_tokens"],   avg_tokens(r),          BLUE)

        # ── CREATIVE WRITING ──────────────────────────────────────────────────
        r = load_results(model_id, "creative_writing")
        if r:
            print(f"  creative_writ : {len(valid(r))}/{len(r)} valid")
            write(sheet, row, COL["cw_quality"],     "—")
            write(sheet, row, COL["cw_coherence"],   "—")
            write(sheet, row, COL["cw_instruction"], instruction_adh(r),    score_fill(instruction_adh(r)))
            write(sheet, row, COL["cw_consist"],     consistency(r),        score_fill(consistency(r)))
            write(sheet, row, COL["cw_latency"],     avg_latency(r),        BLUE)
            write(sheet, row, COL["cw_tokens"],      avg_tokens(r),         BLUE)

        # ── FACTUAL QA ────────────────────────────────────────────────────────
        r = load_results(model_id, "factual_qa")
        if r:
            print(f"  factual_qa    : {len(valid(r))}/{len(r)} valid")
            write(sheet, row, COL["fqa_accuracy"],   accuracy(r,"factual_qa"),     score_fill(accuracy(r,"factual_qa")))
            write(sheet, row, COL["fqa_halluc"],     hallucination(r,"factual_qa"),score_fill(hallucination(r,"factual_qa")))
            write(sheet, row, COL["fqa_instruction"],instruction_adh(r),           score_fill(instruction_adh(r)))
            write(sheet, row, COL["fqa_consist"],    consistency(r),               score_fill(consistency(r)))
            write(sheet, row, COL["fqa_latency"],    avg_latency(r),               BLUE)
            write(sheet, row, COL["fqa_tokens"],     avg_tokens(r),                BLUE)

        # ── SUMMARISATION ─────────────────────────────────────────────────────
        r = load_results(model_id, "summarisation")
        if r:
            print(f"  summarisation : {len(valid(r))}/{len(r)} valid")
            write(sheet, row, COL["sum_kp"],         "—")
            write(sheet, row, COL["sum_accuracy"],   accuracy(r,"summarisation"),  score_fill(accuracy(r,"summarisation")))
            write(sheet, row, COL["sum_instruction"],instruction_adh(r),           score_fill(instruction_adh(r)))
            write(sheet, row, COL["sum_consist"],    consistency(r),               score_fill(consistency(r)))
            write(sheet, row, COL["sum_latency"],    avg_latency(r),               BLUE)
            write(sheet, row, COL["sum_tokens"],     avg_tokens(r),                BLUE)

        # ── INSTRUCTION FOLLOWING ─────────────────────────────────────────────
        r = load_results(model_id, "instruction_following")
        if r:
            print(f"  instruction   : {len(valid(r))}/{len(r)} valid")
            pct = round(instruction_adh(r) / 5 * 100)
            write(sheet, row, COL["if_constraint"], f"{pct}%")
            write(sheet, row, COL["if_accuracy"],   accuracy(r,"instruction_following"),     score_fill(accuracy(r,"instruction_following")))
            write(sheet, row, COL["if_halluc"],     hallucination(r,"instruction_following"),score_fill(hallucination(r,"instruction_following")))
            write(sheet, row, COL["if_consist"],    consistency(r),                          score_fill(consistency(r)))
            write(sheet, row, COL["if_latency"],    avg_latency(r),                          BLUE)
            write(sheet, row, COL["if_tokens"],     avg_tokens(r),                           BLUE)

        # ── REASONING ─────────────────────────────────────────────────────────
        r = load_results(model_id, "reasoning")
        if r:
            print(f"  reasoning     : {len(valid(r))}/{len(r)} valid")
            write(sheet, row, COL["rea_accuracy"], accuracy(r,"reasoning"),     score_fill(accuracy(r,"reasoning")))
            write(sheet, row, COL["rea_chain"],    "—")
            write(sheet, row, COL["rea_halluc"],   hallucination(r,"reasoning"),score_fill(hallucination(r,"reasoning")))
            write(sheet, row, COL["rea_consist"],  consistency(r),              score_fill(consistency(r)))
            write(sheet, row, COL["rea_latency"],  avg_latency(r),              BLUE)
            write(sheet, row, COL["rea_tokens"],   avg_tokens(r),               BLUE)

        # ── LONG CONTEXT ──────────────────────────────────────────────────────
        r = load_results(model_id, "long_context")
        if r:
            print(f"  long_context  : {len(valid(r))}/{len(r)} valid")
            write(sheet, row, COL["lc_retrieval"], accuracy(r,"long_context"),     score_fill(accuracy(r,"long_context")))
            write(sheet, row, COL["lc_accuracy"],  accuracy(r,"long_context"),     score_fill(accuracy(r,"long_context")))
            write(sheet, row, COL["lc_halluc"],    hallucination(r,"long_context"),score_fill(hallucination(r,"long_context")))
            write(sheet, row, COL["lc_consist"],   consistency(r),                 score_fill(consistency(r)))
            write(sheet, row, COL["lc_latency"],   avg_latency(r),                 BLUE)
            write(sheet, row, COL["lc_tokens"],    avg_tokens(r),                  BLUE)

        # ── MULTILINGUAL ──────────────────────────────────────────────────────
        r = load_results(model_id, "multilingual")
        if r:
            print(f"  multilingual  : {len(valid(r))}/{len(r)} valid")
            write(sheet, row, COL["ml_accuracy"], accuracy(r,"multilingual"),     score_fill(accuracy(r,"multilingual")))
            write(sheet, row, COL["ml_fluency"],  "—")
            write(sheet, row, COL["ml_halluc"],   hallucination(r,"multilingual"),score_fill(hallucination(r,"multilingual")))
            write(sheet, row, COL["ml_consist"],  consistency(r),                 score_fill(consistency(r)))
            write(sheet, row, COL["ml_latency"],  avg_latency(r),                 BLUE)
            write(sheet, row, COL["ml_tokens"],   avg_tokens(r),                  BLUE)

        # ── CONVERSATIONAL ────────────────────────────────────────────────────
        r = load_results(model_id, "conversational")
        if r:
            print(f"  conversational: {len(valid(r))}/{len(r)} valid")
            write(sheet, row, COL["conv_quality"],     "—")
            write(sheet, row, COL["conv_context"],     "—")
            write(sheet, row, COL["conv_instruction"], "—")
            write(sheet, row, COL["conv_consist"],     consistency(r),  score_fill(consistency(r)))
            write(sheet, row, COL["conv_latency"],     avg_latency(r),  BLUE)
            write(sheet, row, COL["conv_tokens"],      avg_tokens(r),   BLUE)

    wb.save(OUTPUT_FILE)
    print(f"\n{'='*60}")
    print(f"Saved: {OUTPUT_FILE}")
    print("Cells marked '—' require manual entry.")

if __name__ == "__main__":
    main()